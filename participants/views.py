from rest_framework import viewsets, filters, pagination
from rest_framework.decorators import action
from rest_framework.response import Response
from openpyxl import load_workbook, Workbook
from .models import Participant, Event
from .serializers import ParticipantSerializer, EventSerializer
import time
from django.utils import timezone

# Pagination class with page size set to 100
class ParticipantPagination(pagination.PageNumberPagination):
    page_size = 100

    def get_paginated_response(self, data):
        first_id = data[0]['id'] if data else None
        last_id = data[-1]['id'] if data else None
        return Response({
            'links': {
                'next': self.get_next_link(),
                'previous': self.get_previous_link()
            },
            'count': self.page.paginator.count,
            'results': data,
            'first_id': first_id,
            'last_id': last_id,
        })

class ParticipantViewSet(viewsets.ModelViewSet):
    queryset = Participant.objects.all().order_by('id')
    serializer_class = ParticipantSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['id','name', 'cpf']
    pagination_class = ParticipantPagination
    
    @action(detail=False)
    def count_delivered(self, request):
        participant_count = Participant.objects.filter(delivered=True).count()
        return Response({'count': participant_count})

    @action(detail=False)
    def count_not_delivered(self, request):
        participant_count = Participant.objects.filter(delivered=False).count()
        return Response({'count': participant_count})

    @action(detail=False)
    def generate_report(self, request):
        participants = Participant.objects.all().order_by('id')

        report_data = []
        for participant in participants:
            report_data.append({
                'id': participant.id,
                'chip': participant.chip,
                'name': participant.name,
                'gender': participant.gender,
                'dob': participant.dob,
                'cpf': participant.cpf,
                'course': participant.course,
                'group': participant.group,
                'shirt': participant.shirt,
                'type': participant.type,
                'team': participant.team,
                'nation': participant.nation,
                'medal_record': participant.medal_record,
                'finisher': participant.finisher,
                'fisio': participant.fisio,
                'extra_shirt': participant.extra_shirt,
                'phone': participant.phone,
                'email': participant.email,
                'delivered': participant.delivered,
                'name_received': participant.name_received,
                'updated_at': participant.updated_at.astimezone(timezone.utc).replace(tzinfo=None) if participant.updated_at else None,
            })

        filename = "media/Report.xlsx"
        self.generate_excel_report(report_data, filename)
        
        return Response({'message': 'Success, report generated successfully.', 'file': filename})

    def generate_excel_report(self, data, filename):
        book = Workbook()
        sheet = book.active
        sheet.append([
            'ID', 'CHIP', 'NAME', 'GENDER', 'DOB', 'CPF', 'COURSE','GROUP',
            'SHIRT', 'TYPE', 'TEAM', 'NATION', 'MEDAL RECORD', 'FINISHER', 'FISIO', 'EXTRA SHIRT',
            'PHONE', 'EMAIL', 'DELIVERED', 'NAME RECEIVED', 'UPDATED AT'           
        ])

        for item in data:
            sheet.append([
                item['id'], item['chip'], item['name'], item['gender'], item['dob'], item['cpf'],
                item['course'], item['group'], item['shirt'], item['type'], item['team'], item['nation'], item['medal_record'],
                item['finisher'], item['fisio'], item['extra_shirt'], item['phone'], item['email'], item['delivered'], item['name_received'], item['updated_at']
            ])

        book.save(filename)

    @action(detail=False)
    def create_participants(self, request):
        start_time = time.time()
        file_path = request.data.get('file_path', "media/Entrega_Kits.xlsx")
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active
        
        participants = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                participant = Participant(
                    id=row[0],
                    chip=row[1],
                    name=row[2].upper() if row[2] else None,
                    gender=row[3],
                    dob=row[4],
                    cpf=str(row[5]).zfill(11),
                    course=row[6],
                    group=row[7],
                    shirt=row[8],
                    type=row[9],
                    team=row[10],
                    nation=row[11],
                    medal_record=row[12],
                    finisher=row[13],
                    fisio=row[14],
                    extra_shirt=row[15],
                    phone=row[16],
                    email=row[17],
                    delivered=False,
                    name_received=None,
                    updated_at=None,
                )
                participants.append(participant)
        
        Participant.objects.bulk_create(participants)
        elapsed_time = time.time() - start_time
        print(f"--- {elapsed_time} seconds ---")
        
        return Response({'message': 'Success, all participants added.', 'time_taken': elapsed_time})

class EventViewSet(viewsets.ModelViewSet):
    queryset = Event.objects.all().order_by('id')
    serializer_class = EventSerializer
